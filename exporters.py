"""
Spreadsheet export of an extraction.

WHY: the PDF report is the artifact you forward to a stakeholder or drop in
a vendor file. It is not the artifact you can actually work with. An ELP is
assembled in a spreadsheet, so a licence schedule that can only be read as
a PDF table has to be retyped by hand -- which is both the slowest step and
the one that introduces transcription errors into the numbers everything
else is calculated from.

This produces a three-sheet workbook instead:
  - Contract Terms : every extracted field, with its page and quote, so a
                     reviewer can verify a value without reopening the PDF
  - Line Items     : the licence schedule as real rows and columns, ready to
                     paste into an entitlement working file
  - Risk Register  : flags with severity, source and evidence

Evidence travels with the data on every sheet. A number in an ELP that
can't be traced back to a page and a quote is a number you'll have to
re-verify later anyway.
"""

from __future__ import annotations

import io
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from schema import ContractExtraction, ExtractedField

# Palette mirrors the app and the PDF so the three deliverables read as one
# family rather than three tools that happen to share a pipeline.
INK = "1C2B33"
ACCENT = "2B6E63"
MUTED = "5B6B66"
BORDER_HEX = "D7DED9"
SEVERITY_FILL = {"high": "F5DEDC", "medium": "FAEEDA", "low": "E4EFE7"}

_THIN = Side(style="thin", color=BORDER_HEX)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

FIELD_GROUPS = [
    ("Identification", [
        ("Vendor", "vendor_name"),
        ("Customer", "customer_name"),
        ("Contract title", "contract_title"),
    ]),
    ("Term & Renewal", [
        ("Effective date", "effective_date"),
        ("Term end date", "term_end_date"),
        ("Term length", "term_length"),
        ("Auto-renewal", "auto_renewal"),
        ("Renewal notice period", "renewal_notice_period_days"),
    ]),
    ("Commercial Terms", [
        ("Contract value", "contract_value"),
        ("Pricing model", "pricing_model"),
        ("Price escalation cap", "price_escalation_cap"),
        ("Payment terms", "payment_terms"),
    ]),
    ("License & Entitlement", [
        ("License metric", "license_metric"),
        ("True-up rights", "true_up_rights"),
    ]),
    ("Compliance & Audit", [
        ("Audit rights present", "audit_rights_present"),
        ("Audit notice period", "audit_notice_period_days"),
        ("Audit frequency", "audit_frequency"),
    ]),
    ("Termination", [
        ("Termination for convenience", "termination_for_convenience"),
        ("Termination for cause", "termination_for_cause"),
        ("Early termination fee", "early_termination_fee"),
    ]),
    ("SLA & Exit", [
        ("SLA summary", "sla_summary"),
        ("Data exit / transition period", "data_exit_transition_period"),
    ]),
]


def _header_row(ws, headers: list[str], row: int = 1):
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=ACCENT)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = _BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


# Excel rejects most ASCII control characters outright (openpyxl raises
# IllegalCharacterError), and OCR output is full of them -- stray \x00 and
# \x0b turn up regularly in text recovered from scans and photos. One such
# character anywhere in the batch would abort the whole workbook, losing every
# contract's data because one page was photographed badly. Cells are also
# capped at 32,767 characters, which a long verbatim clause can approach.
_ILLEGAL_XLSX = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_MAX_CELL_CHARS = 32000


def _clean_cell(value):
    """Make any value safe to write to a worksheet cell."""
    if value is None or isinstance(value, (int, float)):
        return value
    text = _ILLEGAL_XLSX.sub("", str(value))
    if len(text) > _MAX_CELL_CHARS:
        text = text[:_MAX_CELL_CHARS] + " […truncated]"
    return text


def _write_cell(ws, row: int, col: int, value, *, bold=False, wrap=True,
                italic=False, color=INK, fill: str | None = None):
    cell = ws.cell(row=row, column=col, value=_clean_cell(value))
    cell.font = Font(bold=bold, italic=italic, color=color, size=10)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
    cell.border = _BORDER
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    return cell


def _set_widths(ws, widths: list[int]):
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _terms_sheet(wb: Workbook, extraction: ContractExtraction, source_filename: str):
    ws = wb.active
    ws.title = "Contract Terms"

    ws["A1"] = "Contract Terms"
    ws["A1"].font = Font(bold=True, size=14, color=INK)
    ws["A2"] = _clean_cell(
        f"Source: {source_filename}  |  Generated {datetime.now():%d %b %Y, %H:%M}  |  "
        "Values are extracted verbatim; blank means not found in the document."
    )
    ws["A2"].font = Font(size=9, italic=True, color=MUTED)

    _header_row(ws, ["Section", "Field", "Value", "Page", "Supporting quote"], row=4)
    _set_widths(ws, [22, 30, 40, 7, 60])

    row = 5
    for section, fields in FIELD_GROUPS:
        first_in_section = True
        for label, attr in fields:
            field: ExtractedField | None = getattr(extraction, attr, None)
            value = field.value if field and field.value else ""
            page = field.evidence.page if field and field.evidence else None
            quote = field.evidence.quote if field and field.evidence else ""

            _write_cell(ws, row, 1, section if first_in_section else "", bold=True)
            _write_cell(ws, row, 2, label)
            _write_cell(ws, row, 3, value or "— not found —",
                        italic=not value, color=INK if value else MUTED)
            _write_cell(ws, row, 4, page if page else "")
            _write_cell(ws, row, 5, quote or "", italic=True, color=MUTED)
            first_in_section = False
            row += 1
    return ws


def _line_items_sheet(wb: Workbook, extraction: ContractExtraction):
    ws = wb.create_sheet("Line Items")
    headers = [
        "Publisher Part #", "Product Name", "License Type", "License Metric",
        "Purchased Rights", "Unit Cost", "Start Date", "End Date",
        "Country", "Page", "Supporting quote",
    ]
    _header_row(ws, headers)
    _set_widths(ws, [18, 34, 14, 18, 18, 14, 13, 13, 14, 7, 46])

    if not extraction.products:
        _write_cell(ws, 2, 1,
                    "No order form, licence schedule or SKU table was found in this document.",
                    italic=True, color=MUTED)
        return ws

    for i, product in enumerate(extraction.products, start=2):
        values = [
            product.publisher_part_number, product.product_name, product.license_type,
            product.license_metric, product.purchased_rights, product.unit_cost,
            product.start_date, product.end_date, product.country_of_agreement,
            product.evidence.page if product.evidence else None,
            product.evidence.quote if product.evidence else None,
        ]
        for col, value in enumerate(values, start=1):
            _write_cell(ws, i, col, value if value else "",
                        italic=(col == 11), color=MUTED if col == 11 else INK)
    return ws


def _risk_sheet(wb: Workbook, extraction: ContractExtraction):
    ws = wb.create_sheet("Risk Register")
    _header_row(ws, ["Severity", "Clause", "Source", "Rule ID", "Finding", "Page", "Supporting quote"])
    _set_widths(ws, [11, 26, 15, 26, 56, 7, 46])

    if not extraction.risk_flags:
        _write_cell(ws, 2, 1, "No risks were flagged for this contract.", italic=True, color=MUTED)
        return ws

    for i, flag in enumerate(extraction.risk_flags, start=2):
        severity = (flag.severity or "").lower()
        _write_cell(ws, i, 1, severity.upper(), bold=True,
                    fill=SEVERITY_FILL.get(severity))
        _write_cell(ws, i, 2, flag.clause)
        _write_cell(ws, i, 3, "Rule-based" if flag.source == "rule" else "AI-suggested")
        _write_cell(ws, i, 4, flag.rule_id or "")
        _write_cell(ws, i, 5, flag.explanation)
        _write_cell(ws, i, 6, flag.evidence.page if flag.evidence and flag.evidence.page else "")
        _write_cell(ws, i, 7, flag.evidence.quote if flag.evidence and flag.evidence.quote else "",
                    italic=True, color=MUTED)
    return ws


def generate_excel_export(extraction: ContractExtraction, source_filename: str) -> bytes:
    """Returns a three-sheet .xlsx workbook as bytes, ready for st.download_button."""
    wb = Workbook()
    _terms_sheet(wb, extraction, source_filename)
    _line_items_sheet(wb, extraction)
    _risk_sheet(wb, extraction)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
