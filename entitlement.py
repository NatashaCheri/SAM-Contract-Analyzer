"""
Multi-contract Excel workbook: the entitlement summary for a manual ELP.

WHAT THIS PRODUCES
------------------
One workbook per batch, with:

  1. Entitlement Summary  -- every licence line item from every contract,
                             stacked into one table, each row carrying the
                             contract and vendor it came from. This is the
                             sheet an ELP is actually built on.
  2. Contracts Overview   -- one row per contract: vendor, term, renewal,
                             audit posture, risk counts, line-item count.
  3. Key Clauses          -- every captured clause, verbatim, grouped by
                             clause type so the same clause can be read
                             across every contract at once.
  4. Risk Register        -- every flag from every contract in one place.
  5. Issues               -- files that couldn't be analysed, and why. Only
                             appears when there are any.
  6. <one sheet per contract> -- the full extracted terms, clauses and flags
                             for that agreement, named after the contract.

WHY THE CONSOLIDATED SHEET COMES FIRST
--------------------------------------
Per-contract sheets are the verification layer; the stacked line-item table
is the working layer. Someone building an ELP lives in the consolidated sheet
and drops into a per-contract sheet only to check a number. Ordering the
workbook the other way round -- contract sheets first, summary buried at the
end -- would put the thing they open every time behind twelve things they
open rarely.

Every consolidated row keeps its source contract, page number and quote, so a
figure that ends up in the ELP can be traced back to a clause without
reopening the original PDF. That traceability is the whole reason to
generate this rather than retype it.
"""

from __future__ import annotations

import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from exporters import FIELD_GROUPS
from risk_rules import risk_summary
from schema import CLAUSE_TYPES

INK = "1C2B33"
ACCENT = "2B6E63"
ACCENT_DARK = "1F5049"
MUTED = "5B6B66"
BORDER_HEX = "D7DED9"
BAND = "F5F8F5"
SEVERITY_FILL = {"high": "F5DEDC", "medium": "FAEEDA", "low": "E4EFE7"}
ISSUE_FILL = "F5DEDC"

_THIN = Side(style="thin", color=BORDER_HEX)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ---------------------------------------------------------------------------
# Sheet naming
# ---------------------------------------------------------------------------

_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def safe_sheet_name(raw_name: str, taken: set[str]) -> str:
    """
    Turn a filename into a legal, unique Excel sheet name.

    Excel's rules are unforgiving and silently break a workbook if violated:
    31 characters maximum, no \\ / * ? : [ ], not blank, and unique within
    the workbook. Contract filenames routinely break all of these at once
    ("Globex - MSA (2026) [signed]/v3.pdf"), so this is not a nicety.

    Truncation is done from the LEFT-hand side of the extension-stripped
    name, keeping the front of the name, because contract filenames are
    usually most distinctive at the start (vendor and agreement type) and
    least distinctive at the end (dates, version tags, "signed", "final").
    """
    stem = re.sub(r"\.(pdf|docx|xlsx|xlsm|csv|txt)$", "", raw_name, flags=re.I)
    cleaned = _INVALID_SHEET_CHARS.sub("-", stem).strip() or "Contract"

    candidate = cleaned[:31]
    if candidate not in taken:
        taken.add(candidate)
        return candidate

    # Collisions are real: "Order Form.pdf" turns up in several vendor folders.
    # Append a counter, trimming the base so the total stays within 31.
    for suffix_num in range(2, 100):
        suffix = f" ({suffix_num})"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        if candidate not in taken:
            taken.add(candidate)
            return candidate

    fallback = f"Contract {len(taken) + 1}"[:31]
    taken.add(fallback)
    return fallback


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------

def _title_block(ws: Worksheet, title: str, subtitle: str):
    ws["A1"] = _clean_cell(title)
    ws["A1"].font = Font(bold=True, size=14, color=INK)
    ws["A2"] = _clean_cell(subtitle)
    ws["A2"].font = Font(size=9, italic=True, color=MUTED)


def _header_row(ws: Worksheet, headers: list[str], row: int):
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=ACCENT)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = _BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


# Excel rejects most ASCII control characters outright (openpyxl raises
# IllegalCharacterError), and OCR output is full of them -- stray \x00 and
# \x0b turn up regularly in text recovered from scans and photos. One such
# character anywhere in the batch would abort the whole workbook, losing every
# contract's data because one page was photographed badly. Cells are also
# capped at 32,767 characters, which a long verbatim clause can approach.
_ILLEGAL_XLSX = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_MAX_CELL_CHARS = 32000


def _clean_cell(value):
    """Make any value safe to write to a worksheet cell."""
    if value is None or isinstance(value, (int, float)):
        return value
    text = _ILLEGAL_XLSX.sub("", str(value))
    if len(text) > _MAX_CELL_CHARS:
        text = text[:_MAX_CELL_CHARS] + " […truncated]"
    return text


def _cell(ws: Worksheet, row: int, col: int, value, *, bold=False, italic=False,
          color=INK, fill: str | None = None, wrap=True):
    cell = ws.cell(row=row, column=col, value=_clean_cell(value))
    cell.font = Font(bold=bold, italic=italic, color=color, size=10)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
    cell.border = _BORDER
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    return cell


def _widths(ws: Worksheet, widths: list[int]):
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _value_of(extraction, attr: str) -> str:
    field = getattr(extraction, attr, None)
    return field.value if field and field.value else ""


# ---------------------------------------------------------------------------
# Sheet 1: Entitlement Summary (the consolidated line-item table)
# ---------------------------------------------------------------------------

ENTITLEMENT_HEADERS = [
    "Contract File", "Vendor", "Customer", "Agreement",
    "Publisher Part #", "Product Name", "License Type", "License Metric",
    "Purchased Rights", "Unit Cost", "Line Start Date", "Line End Date",
    "Country", "Contract Term End", "Page", "Supporting Quote",
]


def _entitlement_sheet(wb: Workbook, batch) -> Worksheet:
    ws = wb.active
    ws.title = "Entitlement Summary"

    contract_count = len(batch.successful)
    _title_block(
        ws, "Entitlement Summary",
        f"All licence line items across {contract_count} contract(s)  |  "
        f"Generated {datetime.now():%d %b %Y, %H:%M}  |  "
        "Each row keeps its source contract, page and quote for traceability.",
    )

    _header_row(ws, ENTITLEMENT_HEADERS, row=4)
    _widths(ws, [26, 22, 20, 26, 18, 32, 13, 18, 16, 13, 14, 14, 12, 16, 7, 44])

    row = 5
    band = False
    for result in batch.successful:
        extraction = result.extraction
        vendor = _value_of(extraction, "vendor_name")
        customer = _value_of(extraction, "customer_name")
        agreement = _value_of(extraction, "contract_title")
        term_end = _value_of(extraction, "term_end_date")
        fill = BAND if band else None

        if not extraction.products:
            # A contract with no line-item table still belongs in the summary.
            # Dropping it would make the sheet look complete while quietly
            # omitting an agreement someone has to account for -- the reader
            # needs to see that it was analysed and had nothing to extract.
            _cell(ws, row, 1, result.filename, bold=True, fill=fill)
            _cell(ws, row, 2, vendor, fill=fill)
            _cell(ws, row, 3, customer, fill=fill)
            _cell(ws, row, 4, agreement, fill=fill)
            _cell(ws, row, 5, "— no licence schedule found in this document —",
                  italic=True, color=MUTED, fill=fill)
            for col in range(6, len(ENTITLEMENT_HEADERS) + 1):
                _cell(ws, row, col, "", fill=fill)
            _cell(ws, row, 14, term_end, fill=fill)
            row += 1
            band = not band
            continue

        for product in extraction.products:
            values = [
                result.filename, vendor, customer, agreement,
                product.publisher_part_number, product.product_name,
                product.license_type, product.license_metric,
                product.purchased_rights, product.unit_cost,
                product.start_date, product.end_date,
                product.country_of_agreement, term_end,
                product.evidence.page if product.evidence else None,
                product.evidence.quote if product.evidence else None,
            ]
            for col, value in enumerate(values, start=1):
                _cell(
                    ws, row, col, value if value else "",
                    bold=(col == 1),
                    italic=(col == 16),
                    color=MUTED if col == 16 else INK,
                    fill=fill,
                )
            row += 1
        band = not band

    if row == 5:
        _cell(ws, 5, 1, "No line items were extracted from any contract in this batch.",
              italic=True, color=MUTED)

    ws.auto_filter.ref = f"A4:{get_column_letter(len(ENTITLEMENT_HEADERS))}{max(row - 1, 5)}"
    return ws


# ---------------------------------------------------------------------------
# Sheet 2: Contracts Overview
# ---------------------------------------------------------------------------

OVERVIEW_HEADERS = [
    "Contract File", "Sheet", "Vendor", "Customer", "Agreement",
    "Effective Date", "Term End", "Auto-Renewal", "Renewal Notice",
    "Contract Value", "License Metric", "Audit Rights", "Audit Notice",
    "Line Items", "High", "Medium", "Low", "Risk Posture", "Pages",
]


def _overview_sheet(wb: Workbook, batch, sheet_names: list[str]) -> Worksheet:
    ws = wb.create_sheet("Contracts Overview")
    _title_block(
        ws, "Contracts Overview",
        "One row per contract. The Sheet column names the tab holding that "
        "contract's full extracted terms.",
    )
    _header_row(ws, OVERVIEW_HEADERS, row=4)
    _widths(ws, [26, 20, 20, 18, 24, 13, 13, 12, 14, 14, 16, 12, 13, 10, 7, 8, 7, 16, 7])

    row = 5
    for result, sheet_name in zip(batch.successful, sheet_names):
        extraction = result.extraction
        stats = risk_summary(extraction.risk_flags)
        counts = stats["counts"]
        values = [
            result.filename,
            sheet_name,
            _value_of(extraction, "vendor_name"),
            _value_of(extraction, "customer_name"),
            _value_of(extraction, "contract_title"),
            _value_of(extraction, "effective_date"),
            _value_of(extraction, "term_end_date"),
            _value_of(extraction, "auto_renewal"),
            _value_of(extraction, "renewal_notice_period_days"),
            _value_of(extraction, "contract_value"),
            _value_of(extraction, "license_metric"),
            _value_of(extraction, "audit_rights_present"),
            _value_of(extraction, "audit_notice_period_days"),
            len(extraction.products),
            counts["high"], counts["medium"], counts["low"],
            stats["posture"], result.pages,
        ]
        for col, value in enumerate(values, start=1):
            fill = None
            if col == 15 and counts["high"]:
                fill = SEVERITY_FILL["high"]
            elif col == 16 and counts["medium"]:
                fill = SEVERITY_FILL["medium"]
            _cell(ws, row, col, value if value not in (None, "") else "",
                  bold=(col == 1), fill=fill)
        row += 1

    ws.auto_filter.ref = f"A4:{get_column_letter(len(OVERVIEW_HEADERS))}{max(row - 1, 5)}"
    return ws


# ---------------------------------------------------------------------------
# Sheet 3: Risk Register (all contracts)
# ---------------------------------------------------------------------------

def _risk_register_sheet(wb: Workbook, batch) -> Worksheet:
    ws = wb.create_sheet("Risk Register")
    _title_block(
        ws, "Risk Register",
        "Every flag from every contract. Rule-based rows are deterministic; "
        "AI-suggested rows need a human read.",
    )
    _header_row(ws, ["Contract File", "Vendor", "Severity", "Clause", "Source",
                     "Rule ID", "Finding", "Page", "Supporting Quote"], row=4)
    _widths(ws, [26, 20, 11, 24, 15, 26, 52, 7, 44])

    row = 5
    for result in batch.successful:
        for flag in result.extraction.risk_flags:
            severity = (flag.severity or "").lower()
            _cell(ws, row, 1, result.filename, bold=True)
            _cell(ws, row, 2, result.vendor)
            _cell(ws, row, 3, severity.upper(), bold=True, fill=SEVERITY_FILL.get(severity))
            _cell(ws, row, 4, flag.clause)
            _cell(ws, row, 5, "Rule-based" if flag.source == "rule" else "AI-suggested")
            _cell(ws, row, 6, flag.rule_id or "")
            _cell(ws, row, 7, flag.explanation)
            _cell(ws, row, 8, flag.evidence.page if flag.evidence and flag.evidence.page else "")
            _cell(ws, row, 9, flag.evidence.quote if flag.evidence and flag.evidence.quote else "",
                  italic=True, color=MUTED)
            row += 1

    if row == 5:
        _cell(ws, 5, 1, "No risks were flagged across this batch.", italic=True, color=MUTED)

    ws.auto_filter.ref = f"A4:I{max(row - 1, 5)}"
    return ws


# ---------------------------------------------------------------------------
# Sheet 4: Key Clauses (the clause finder, as a filterable table)
# ---------------------------------------------------------------------------

def _clauses_sheet(wb: Workbook, batch) -> Worksheet:
    """
    Every captured clause across the batch, one row each.

    Ordered by CLAUSE TYPE first and contract second, deliberately. Someone
    using this sheet is asking "what does everything say about audit rights?",
    not "what's in the Globex contract?" -- that second question is what the
    per-contract tabs are for. Grouping by type puts all twelve audit clauses
    next to each other so they can be read down the column, and an autofilter
    on top narrows to one type in a click.
    """
    ws = wb.create_sheet("Key Clauses")
    _title_block(
        ws, "Key Clauses",
        "Clause text is copied verbatim from each contract. The plain-English column "
        "is a reading of it, not contract language. Filter by Clause Type to compare "
        "the same clause across every agreement.",
    )
    _header_row(ws, ["Clause Type", "Contract File", "Vendor", "Heading",
                     "Clause Text (verbatim)", "Plain English", "Page"], row=4)
    _widths(ws, [26, 26, 20, 26, 84, 46, 7])

    order = {name: i for i, name in enumerate(CLAUSE_TYPES)}
    rows = [
        (clause, result)
        for result in batch.successful
        for clause in result.extraction.key_clauses
    ]
    rows.sort(key=lambda pair: (order.get(pair[0].clause_type, 999), pair[1].filename))

    row_index = 5
    previous_type = None
    band = False
    for clause, result in rows:
        if clause.clause_type != previous_type:
            band = not band
            previous_type = clause.clause_type
        fill = BAND if band else None

        _cell(ws, row_index, 1, clause.clause_type, bold=True, fill=fill)
        _cell(ws, row_index, 2, result.filename, fill=fill)
        _cell(ws, row_index, 3, result.vendor, fill=fill)
        _cell(ws, row_index, 4, clause.heading or "", fill=fill)
        _cell(ws, row_index, 5, clause.text or "", fill=fill)
        _cell(ws, row_index, 6, clause.summary or "", italic=True, color=MUTED, fill=fill)
        _cell(ws, row_index, 7, clause.page if clause.page else "", fill=fill)
        ws.row_dimensions[row_index].height = 58
        row_index += 1

    if row_index == 5:
        _cell(ws, 5, 1, "No clauses were captured across this batch.",
              italic=True, color=MUTED)

    ws.auto_filter.ref = f"A4:G{max(row_index - 1, 5)}"
    return ws


# ---------------------------------------------------------------------------
# Sheet 5: Issues (only when something failed)
# ---------------------------------------------------------------------------

def _issues_sheet(wb: Workbook, batch) -> Worksheet | None:
    if not batch.failed:
        return None

    ws = wb.create_sheet("Issues")
    _title_block(
        ws, "Files That Could Not Be Analysed",
        "These documents are NOT represented in the entitlement summary. "
        "Resolve and re-run them before treating the summary as complete.",
    )
    _header_row(ws, ["File", "What went wrong"], row=4)
    _widths(ws, [34, 90])

    for i, result in enumerate(batch.failed, start=5):
        _cell(ws, i, 1, result.filename, bold=True, fill=ISSUE_FILL)
        _cell(ws, i, 2, result.error or "Unknown error")
    return ws


# ---------------------------------------------------------------------------
# Per-contract detail sheets
# ---------------------------------------------------------------------------

def _contract_sheet(wb: Workbook, result, sheet_name: str) -> Worksheet:
    ws = wb.create_sheet(sheet_name)
    extraction = result.extraction

    _title_block(
        ws, result.vendor,
        f"Source: {result.filename}  |  {result.pages} page(s)"
        + (f", {result.ocr_pages} via OCR" if result.ocr_pages else "")
        + "  |  Blank value means the term was not found in this document.",
    )

    _header_row(ws, ["Section", "Field", "Value", "Page", "Supporting quote"], row=4)
    _widths(ws, [22, 30, 40, 7, 58])

    row = 5
    for section, fields in FIELD_GROUPS:
        first = True
        for label, attr in fields:
            field = getattr(extraction, attr, None)
            value = field.value if field and field.value else ""
            page = field.evidence.page if field and field.evidence else None
            quote = field.evidence.quote if field and field.evidence else ""

            _cell(ws, row, 1, section if first else "", bold=True)
            _cell(ws, row, 2, label)
            _cell(ws, row, 3, value or "— not found —",
                  italic=not value, color=INK if value else MUTED)
            _cell(ws, row, 4, page if page else "")
            _cell(ws, row, 5, quote or "", italic=True, color=MUTED)
            first = False
            row += 1

    # Line items for this contract, so the sheet stands alone as the
    # verification view for anything in the consolidated summary.
    row += 1
    _cell(ws, row, 1, "Licence line items", bold=True, color=ACCENT_DARK)
    row += 1
    line_headers = ["Part #", "Product Name", "License Type", "License Metric", "Purchased Rights"]
    for col, text in enumerate(line_headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=ACCENT)
        cell.border = _BORDER
    row += 1

    if extraction.products:
        for product in extraction.products:
            for col, value in enumerate([
                product.publisher_part_number, product.product_name,
                product.license_type, product.license_metric, product.purchased_rights,
            ], start=1):
                _cell(ws, row, col, value or "")
            row += 1
    else:
        _cell(ws, row, 1, "No licence schedule found in this document.",
              italic=True, color=MUTED)
        row += 1

    # Key clauses for this contract, so the sheet works standalone as the
    # single place to review one agreement.
    row += 1
    _cell(ws, row, 1, "Key clauses", bold=True, color=ACCENT_DARK)
    row += 1
    for col, text in enumerate(["Clause Type", "Heading", "Clause Text (verbatim)", "Plain English", "Page"], start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=ACCENT)
        cell.border = _BORDER
    row += 1

    if extraction.key_clauses:
        for clause in extraction.key_clauses:
            _cell(ws, row, 1, clause.clause_type, bold=True)
            _cell(ws, row, 2, clause.heading or "")
            _cell(ws, row, 3, clause.text or "")
            _cell(ws, row, 4, clause.summary or "", italic=True, color=MUTED)
            _cell(ws, row, 5, clause.page if clause.page else "")
            ws.row_dimensions[row].height = 58
            row += 1
    else:
        _cell(ws, row, 1, "No clauses captured from this document.", italic=True, color=MUTED)
        row += 1

    # Risk flags for this contract
    row += 1
    _cell(ws, row, 1, "Risk flags", bold=True, color=ACCENT_DARK)
    row += 1
    for col, text in enumerate(["Severity", "Clause", "Source", "Finding"], start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=ACCENT)
        cell.border = _BORDER
    row += 1

    if extraction.risk_flags:
        for flag in extraction.risk_flags:
            severity = (flag.severity or "").lower()
            _cell(ws, row, 1, severity.upper(), bold=True, fill=SEVERITY_FILL.get(severity))
            _cell(ws, row, 2, flag.clause)
            _cell(ws, row, 3, "Rule-based" if flag.source == "rule" else "AI-suggested")
            _cell(ws, row, 4, flag.explanation)
            row += 1
    else:
        _cell(ws, row, 1, "No risks flagged.", italic=True, color=MUTED)

    return ws


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_batch_workbook(batch) -> bytes:
    """
    Build the full multi-contract workbook and return it as bytes.

    Takes a batch.BatchResult. Sheet order is deliberate: the consolidated
    entitlement table first, then the roll-ups, then one detail sheet per
    contract.
    """
    import io

    wb = Workbook()

    _entitlement_sheet(wb, batch)

    # Sheet names are resolved before the overview is written, so the overview
    # can point each contract row at the tab holding its detail.
    #
    # Built as a LIST positionally aligned to batch.successful, not a dict
    # keyed by filename: two genuinely different contracts routinely share a
    # filename ("Order Form.pdf" appears in most vendor folders), and keying
    # by name silently collapses them onto one entry -- the first contract
    # loses its sheet reference and openpyxl quietly renames the duplicate
    # tab behind your back.
    taken: set[str] = {
        "Entitlement Summary", "Contracts Overview", "Key Clauses",
        "Risk Register", "Issues",
    }
    sheet_names = [safe_sheet_name(r.filename, taken) for r in batch.successful]

    _overview_sheet(wb, batch, sheet_names)
    _clauses_sheet(wb, batch)
    _risk_register_sheet(wb, batch)
    _issues_sheet(wb, batch)

    for result, sheet_name in zip(batch.successful, sheet_names):
        _contract_sheet(wb, result, sheet_name)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
